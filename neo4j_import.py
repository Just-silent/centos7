# coding:UTF-8
# author    :Just_silent
# init time :2021/7/6 15:10
# file      :neo4j_import.py
# IDE       :PyCharm

import openpyxl
from tqdm import tqdm
from py2neo import Graph, Node, Relationship, NodeMatcher, RelationshipMatcher, Subgraph


class Neo4jDictImport():
    def __init__(self):
        self.graf = Graph(
            "http://192.168.8.115/:7474/",
            user="neo4j",
            password="123456"
        )
        self.node_match = NodeMatcher(self.graf)
        self.rel_match = RelationshipMatcher(self.graf)

    def _create_node(self, root, node):
        rels = ['s_d1', 'd1_d2', 'd2_d3', 'd3_d4', 'd4_k', 'k_q', 'q_a']
        node_names = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h']
        if root==[]:
            # 直接创建系统节点
            sql = 'match (n:' + node[0] + '{name:\'' + node[1] + '\'}) return n'
            result = self.graf.run(sql).data()
            if result==[]:
                sql = 'create (n:' + '{}'.format(node[0])+'{name:\''+ node[1] +'\'})'
                self.graf.run(sql)
        else:
            # 获取root部分末尾节点，并穿件关系和新的节点
            # 获取末尾节点id
            i=0
            sql_last_of_root = 'match ({}:'.format(node_names[0])+root[0][0]+'{name:\''+root[0][1]+'\'})'
            for i in range(1, len(root)):
                if root[i][1] is not None:
                    sql_last_of_root += '-['+rels[i-1]+']->({}:'.format(node_names[i])+root[i][0]+'{name:\''+root[i][1]+'\'})'
            sql_last_of_root+='return {}'.format(node_names[i])
            last_of_root = self.graf.run(sql_last_of_root).data()[0][node_names[i]]
            # 判断node是否在root下
            sql_next_of_root = 'match (m)-[{}]->(n) where id(m)={} return n'.format(rels[i], last_of_root.identity)
            next_of_root = self.graf.run(sql_next_of_root).data()
            if next_of_root==[] or node[1] not in [x['n']['name'] for x in next_of_root]:
                sql_create_node = 'create (m:'+node[0]+'{name:\"'+node[1].replace('\n', '').replace('\\', '')+'\"}) return m'
                new_node = self.graf.run(sql_create_node).data()[0]['m']
                sql = 'match (n),(m) where id(n)={} and id(m)={} create (n)-[r:{}]->(m) return m'.format(last_of_root.identity, new_node.identity, rels[i])
                self.graf.run(sql)
        pass

    def _create_line(self, element):
        keys = list(element.keys())
        root = []
        for i in range(len(keys)):
            if element[keys[i]] is not None:
                node = (keys[i], element[keys[i]])
                self._create_node(root, node)
                root.append(node)
        pass

    def create_neo4j(self, path, sheet='neo'):
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet]
        maxrow = ws.max_row  # 最大行
        maxcol = ws.max_column  # 最大列
        names = [ws.cell(1, i).value for i in range(1, maxcol + 1)]
        for line in tqdm(range(2, maxrow + 1)):
            element = {
                'system': ws.cell(line, names.index('system') + 1).value if 'system' in names else None,
                'domain1': ws.cell(line, names.index('first') + 1).value if 'first' in names else None,
                'domain2': ws.cell(line, names.index('second') + 1).value if 'second' in names else None,
                'domain3': ws.cell(line, names.index('third') + 1).value if 'third' in names else None,
                'domain4': ws.cell(line, names.index('fourth') + 1).value if 'fourth' in names else None,
                'keyword': ws.cell(line, names.index('keyword') + 1).value if 'keyword' in names else None,
                'question': ws.cell(line, names.index('question') + 1).value if 'question' in names else None,
                'operate': ws.cell(line, names.index('operate') + 1).value if 'operate' in names else None,
            }
            self._create_line(element)



if __name__ == '__main__':
    xlsx_path = 'xlsx_path'
    neo4j_dict_import = Neo4jDictImport()
    neo4j_dict_import.create_neo4j(xlsx_path)